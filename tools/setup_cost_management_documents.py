from pathlib import Path
from datetime import datetime
import shutil

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
SEED_DIR = DATA_DIR / "seeds"
BACKUP_DIR = DATA_DIR / "backups"
TEMPLATE_DIR = BASE_DIR / "templates" / "documents"

DOCUMENT_FILE_NAMES = [
    DATA_DIR / "document_master.xlsx",
    SEED_DIR / "document_master_default.xlsx",
]

QUESTIONNAIRE_FILE_NAMES = [
    DATA_DIR / "questionnaire_data.xlsx",
    SEED_DIR / "questionnaire_data_default.xlsx",
]


COST_TEMPLATES = {
    "COST-001": {
        "filename": "COST-001_product_project_cost_management.xlsx",
        "document_name": "製品別・案件別原価管理表",
        "description": "製品別・案件別に材料費、外注費、労務費、製造間接費を把握するための管理表。",
        "owner_department": "経理部",
        "owner": "経理部",
        "importance": "高",
        "risk_level": "高",
        "headers": [
            "対象月",
            "製品名・案件名",
            "売上高",
            "材料費",
            "外注費",
            "労務費",
            "製造間接費",
            "総原価",
            "粗利",
            "粗利率",
            "備考",
        ],
        "sample_rows": [
            ["2026-05", "製品A", "", "", "", "", "", "", "", "", ""],
        ],
    },
    "COST-002": {
        "filename": "COST-002_cost_classification_rule.xlsx",
        "document_name": "原価分類ルール表",
        "description": "材料費、外注費、労務費、製造間接費の分類ルールを定義するための表。",
        "owner_department": "経理部",
        "owner": "経理部",
        "importance": "高",
        "risk_level": "高",
        "headers": [
            "分類",
            "対象となる費用",
            "具体例",
            "集計方法",
            "入力担当",
            "確認担当",
            "備考",
        ],
        "sample_rows": [
            ["材料費", "製品に直接使用する材料", "樹脂、金属部材、添加剤など", "", "", "", ""],
            ["外注費", "外部委託した加工・検査費用", "外注加工、外部検査など", "", "", "", ""],
            ["労務費", "製造に関わる人件費", "製造担当者の作業時間など", "", "", "", ""],
            ["製造間接費", "直接紐づけにくい製造関連費用", "電力、設備償却、補助材料など", "", "", "", ""],
        ],
    },
    "COST-003": {
        "filename": "COST-003_monthly_cost_gross_profit_check.xlsx",
        "document_name": "月次原価率・粗利率確認表",
        "description": "月次で原価率、粗利率、前月差異、主な変動要因を確認するための表。",
        "owner_department": "経理部",
        "owner": "経理部",
        "importance": "高",
        "risk_level": "高",
        "headers": [
            "対象月",
            "売上高",
            "総原価",
            "原価率",
            "粗利",
            "粗利率",
            "前月粗利率",
            "差異",
            "主な要因",
            "対応方針",
            "確認者",
        ],
        "sample_rows": [
            ["2026-05", "", "", "", "", "", "", "", "", "", ""],
        ],
    },
    "COST-004": {
        "filename": "COST-004_estimated_actual_cost_comparison.xlsx",
        "document_name": "見積原価・実績原価比較表",
        "description": "見積時の原価と実績原価を比較し、差異理由を確認するための表。",
        "owner_department": "経理部",
        "owner": "製造部",
        "importance": "中",
        "risk_level": "中",
        "headers": [
            "案件名",
            "製品名",
            "見積原価",
            "実績原価",
            "差異",
            "差異率",
            "差異理由",
            "再発防止・改善策",
            "担当者",
            "確認日",
        ],
        "sample_rows": [
            ["案件A", "製品A", "", "", "", "", "", "", "", ""],
        ],
    },
    "COST-005": {
        "filename": "COST-005_defect_rework_yield_loss_management.xlsx",
        "document_name": "不良・手直し・歩留まりロス管理表",
        "description": "不良、手直し、歩留まりロスが原価に与える影響を把握するための管理表。",
        "owner_department": "製造部",
        "owner": "製造部",
        "importance": "中",
        "risk_level": "中",
        "headers": [
            "発生日",
            "製品名",
            "不良内容",
            "不良数量",
            "手直し時間",
            "廃棄数量",
            "歩留まりロス金額",
            "原因",
            "対策",
            "担当者",
            "確認者",
        ],
        "sample_rows": [
            ["", "製品A", "", "", "", "", "", "", "", "", ""],
        ],
    },
}


QUESTION_LINKS = {
    "Q-006": "COST-001",
    "Q-007": "COST-002",
    "Q-008": "COST-003",
    "Q-009": "COST-004",
    "Q-010": "COST-005",
}


def ensure_directories():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SEED_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


def timestamp_text():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def backup_file(path):
    if not path.exists():
        return None

    backup_path = BACKUP_DIR / f"{path.stem}_{timestamp_text()}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def normalize_df(df):
    df = df.fillna("")

    for column in df.columns:
        df[column] = df[column].astype(str)

    return df


def read_all_sheets(excel_path):
    try:
        sheets = pd.read_excel(excel_path, sheet_name=None, dtype=str)
    except Exception:
        return {}

    return {
        sheet_name: normalize_df(df)
        for sheet_name, df in sheets.items()
    }


def write_all_sheets(excel_path, sheets):
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        for sheet_name, df in sheets.items():
            normalize_df(df).to_excel(writer, sheet_name=sheet_name, index=False)


def make_relative_path(path):
    return str(path.resolve().relative_to(BASE_DIR.resolve())).replace("\\", "/")


def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F2937")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="B7C0CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A4"


def create_cost_template_file(document_id, template_info):
    file_path = TEMPLATE_DIR / template_info["filename"]

    if file_path.exists():
        return file_path

    wb = Workbook()
    ws = wb.active
    ws.title = "入力シート"

    headers = template_info["headers"]

    ws.cell(row=1, column=1, value=template_info["document_name"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws.cell(row=2, column=1, value="目的")
    ws.cell(row=2, column=2, value=template_info["description"])
    if len(headers) >= 2:
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))

    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_index, value=header)

    for row_index, sample_row in enumerate(template_info["sample_rows"], start=4):
        for col_index, value in enumerate(sample_row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    for col_index in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=3, column=col_index).column_letter].width = 18

    style_worksheet(ws)

    wb.save(file_path)

    return file_path


def create_all_template_files():
    print("=== 原価管理 Excel ひな形作成 ===")

    created_paths = {}

    for document_id, template_info in COST_TEMPLATES.items():
        file_path = create_cost_template_file(document_id, template_info)
        created_paths[document_id] = file_path
        print(f"[OK] {document_id}: {file_path}")

    return created_paths


def update_document_master(excel_path, template_paths):
    if not excel_path.exists():
        print(f"[SKIP] 文書台帳ファイルがありません: {excel_path}")
        return False

    sheets = read_all_sheets(excel_path)

    if "documents" not in sheets:
        print(f"[SKIP] documents シートがありません: {excel_path}")
        return False

    df = sheets["documents"]

    if "id" not in df.columns:
        print(f"[SKIP] id 列がありません: {excel_path}")
        return False

    required_columns = [
        "id",
        "category",
        "document_name",
        "description",
        "owner_department",
        "owner",
        "importance",
        "risk_level",
        "status",
        "template_file_path",
        "completed_file_path",
        "completed_file_name",
        "completed_at",
        "completed_by",
        "established_date",
        "revised_date",
        "next_review_date",
        "created_at",
        "updated_at",
    ]

    for column in required_columns:
        if column not in df.columns:
            df[column] = ""

    existing_ids = set(df["id"].astype(str).tolist())
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_rows = []
    updated_count = 0

    for document_id, template_info in COST_TEMPLATES.items():
        template_file_path = make_relative_path(template_paths[document_id])

        if document_id in existing_ids:
            target_index = df.index[df["id"].astype(str) == document_id].tolist()
            if not target_index:
                continue

            index = target_index[0]

            update_map = {
                "category": "原価管理",
                "document_name": template_info["document_name"],
                "description": template_info["description"],
                "owner_department": template_info["owner_department"],
                "owner": template_info["owner"],
                "importance": template_info["importance"],
                "risk_level": template_info["risk_level"],
                "status": df.loc[index, "status"] or "未整備",
                "template_file_path": template_file_path,
                "updated_at": now_text,
            }

            for key, value in update_map.items():
                if key in df.columns and str(df.loc[index, key]).strip() != str(value):
                    df.loc[index, key] = value
                    updated_count += 1

            continue

        new_row = {}

        for column in df.columns:
            new_row[column] = ""

        new_row.update({
            "id": document_id,
            "category": "原価管理",
            "document_name": template_info["document_name"],
            "description": template_info["description"],
            "owner_department": template_info["owner_department"],
            "owner": template_info["owner"],
            "importance": template_info["importance"],
            "risk_level": template_info["risk_level"],
            "status": "未整備",
            "template_file_path": template_file_path,
            "created_at": now_text,
            "updated_at": now_text,
        })

        new_rows.append(new_row)

    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    if not new_rows and updated_count == 0:
        print(f"[OK] 文書台帳は更新不要です: {excel_path}")
        return True

    backup_path = backup_file(excel_path)
    print(f"[BACKUP] {excel_path.name}: {backup_path}")

    sheets["documents"] = df
    write_all_sheets(excel_path, sheets)

    print(
        f"[OK] 文書台帳を更新しました: {excel_path} "
        f"追加 {len(new_rows)} 件 / 更新 {updated_count} 箇所"
    )
    return True


def update_questionnaire_links(excel_path):
    if not excel_path.exists():
        print(f"[SKIP] 質問票ファイルがありません: {excel_path}")
        return False

    sheets = read_all_sheets(excel_path)

    if "questions" not in sheets:
        print(f"[SKIP] questions シートがありません: {excel_path}")
        return False

    df = sheets["questions"]

    if "id" not in df.columns:
        print(f"[SKIP] id 列がありません: {excel_path}")
        return False

    required_columns = [
        "recommended_due_date_days",
        "related_document_id",
        "related_app",
        "related_item_type",
    ]

    for column in required_columns:
        if column not in df.columns:
            df[column] = ""

    updated_count = 0

    for question_id, document_id in QUESTION_LINKS.items():
        target_index = df.index[df["id"].astype(str) == question_id].tolist()

        if not target_index:
            continue

        index = target_index[0]

        updates = {
            "related_document_id": document_id,
            "related_app": "documents",
            "related_item_type": "cost_document",
        }

        for key, value in updates.items():
            if str(df.loc[index, key]).strip() != value:
                df.loc[index, key] = value
                updated_count += 1

        if "recommended_due_date" in df.columns:
            old_value = str(df.loc[index, "recommended_due_date"]).strip()
            if old_value and not str(df.loc[index, "recommended_due_date_days"]).strip():
                df.loc[index, "recommended_due_date_days"] = old_value
                updated_count += 1

    if updated_count == 0:
        print(f"[OK] 質問票リンクは更新不要です: {excel_path}")
        return True

    backup_path = backup_file(excel_path)
    print(f"[BACKUP] {excel_path.name}: {backup_path}")

    sheets["questions"] = df
    write_all_sheets(excel_path, sheets)

    print(f"[OK] 質問票の関連文書リンクを {updated_count} 箇所更新しました: {excel_path}")
    return True


def main():
    ensure_directories()

    print("=== 原価管理 文書台帳・質問票リンク・ひな形 セットアップ開始 ===")

    template_paths = create_all_template_files()

    for document_path in DOCUMENT_FILE_NAMES:
        update_document_master(document_path, template_paths)

    for questionnaire_path in QUESTIONNAIRE_FILE_NAMES:
        update_questionnaire_links(questionnaire_path)

    print("=== 原価管理 文書台帳・質問票リンク・ひな形 セットアップ完了 ===")


if __name__ == "__main__":
    main()